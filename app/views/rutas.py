from flask import Flask, render_template, request, redirect, url_for, session
from flask_mail import Mail, Message
from app.controllers.usuario_facade import UsuarioFacade
from config.database import MAIL_CONFIG
import secrets

app = Flask(__name__, template_folder='templates')
app.secret_key = 'gestion_vuelos_secret_2024'

# Configurar correo
for key, value in MAIL_CONFIG.items():
    app.config[key] = value
mail = Mail(app)

facade = UsuarioFacade()

# Tokens temporales para recuperación
tokens_recuperacion = {}

# PROTECCIÓN DE RUTAS
def login_requerido():
    return 'usuario_id' not in session

def solo_admin():
    return session.get('rol') != 'admin'


# LOGIN
@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'usuario_id' in session:
        return redirect(url_for('index'))

    if request.method == 'GET':
        return render_template('login.html', error=None, correo='')

    correo   = request.form.get('correo', '')
    password = request.form.get('password', '')

    resultado = facade.login(correo, password)

    if not resultado['exito']:
        return render_template('login.html', error=resultado['error'], correo=correo)

    usuario = resultado['usuario']
    session['usuario_id'] = usuario['id']
    session['nombre']     = usuario['nombre']
    session['rol']        = usuario['rol']
    session['permisos']   = usuario['permisos']

    return redirect(url_for('index'))


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# REGISTRO
@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'GET':
        roles = facade.obtener_roles()
        return render_template('registro.html', errores=[], roles=roles)

    datos = {
        'nombre':   request.form.get('nombre', ''),
        'correo':   request.form.get('correo', ''),
        'password': request.form.get('password', ''),
        'rol_id':   '3'  # cliente por defecto
    }

    resultado = facade.crear_usuario(datos)

    if not resultado['exito']:
        roles = facade.obtener_roles()
        return render_template('registro.html', errores=resultado['errores'], roles=roles)

    # Enviar correo de bienvenida
    try:
        msg = Message(
            subject='Bienvenido a Gestión de Vuelos',
            recipients=[datos['correo']]
        )
        msg.html = f"""
        <div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto;background:#1a1a1a;color:#e0e0e0;padding:32px;border-radius:12px;">
            <h1 style="color:#ffffff;">Bienvenido a Gestión de Vuelos</h1>
            <p>Hola <strong>{datos['nombre']}</strong>,</p>
            <p>Tu cuenta ha sido creada exitosamente.</p>
            <p><strong>Correo:</strong> {datos['correo']}</p>
            <p>Ya puedes iniciar sesión en el sistema.</p>
            <a href="http://localhost:5000/login" style="background:#444;color:#fff;padding:10px 24px;border-radius:7px;text-decoration:none;">Iniciar Sesión</a>
        </div>
        """
        mail.send(msg)
    except Exception as e:
        print(f'Error enviando correo: {e}')

    return redirect(url_for('login'))


# RECUPERAR CONTRASEÑA
@app.route('/recuperar', methods=['GET', 'POST'])
def recuperar():
    if request.method == 'GET':
        return render_template('recuperar.html', mensaje=None, error=None)

    correo = request.form.get('correo', '')
    from app.models.usuario import Usuario
    repo = Usuario()
    usuario = repo.obtener_por_correo(correo)

    if not usuario:
        return render_template('recuperar.html',
                               error='El correo no está registrado.',
                               mensaje=None)

    token = secrets.token_urlsafe(32)
    tokens_recuperacion[token] = correo

    try:
        msg = Message(
            subject='Recuperación de contraseña — Gestión de Vuelos',
            recipients=[correo]
        )
        msg.html = f"""
        <div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto;background:#1a1a1a;color:#e0e0e0;padding:32px;border-radius:12px;">
            <h1 style="color:#ffffff;"> Recuperar Contraseña</h1>
            <p>Hola <strong>{usuario['nombre']}</strong>,</p>
            <p>Recibimos una solicitud para restablecer tu contraseña.</p>
            <p>Haz clic en el botón para crear una nueva contraseña:</p>
            <a href="http://localhost:5000/restablecer/{token}" style="background:#444;color:#fff;padding:10px 24px;border-radius:7px;text-decoration:none;">Restablecer Contraseña</a>
            <p style="margin-top:16px;color:#aaa;">Si no solicitaste esto, ignora este correo.</p>
        </div>
        """
        mail.send(msg)
    except Exception as e:
        print(f'Error enviando correo: {e}')

    return render_template('recuperar.html',
                           mensaje='Te enviamos un correo con instrucciones.',
                           error=None)


@app.route('/restablecer/<token>', methods=['GET', 'POST'])
def restablecer(token):
    correo = tokens_recuperacion.get(token)
    if not correo:
        return render_template('restablecer.html', error='Token inválido o expirado.', token=token)

    if request.method == 'GET':
        return render_template('restablecer.html', error=None, token=token)

    password = request.form.get('password', '')
    confirmar = request.form.get('confirmar', '')

    if len(password) < 6:
        return render_template('restablecer.html',
                               error='La contraseña debe tener al menos 6 caracteres.', token=token)
    if password != confirmar:
        return render_template('restablecer.html',
                               error='Las contraseñas no coinciden.', token=token)

    import bcrypt
    from app.models.usuario import Usuario
    password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    repo = Usuario()
    repo.actualizar_password(correo, password_hash)
    tokens_recuperacion.pop(token, None)

    return redirect(url_for('login'))


# RUTAS PROTEGIDAS
@app.route('/')
def index():
    if login_requerido():
        return redirect(url_for('login'))
    return redirect(url_for('listar'))


@app.route('/usuarios')
def listar():
    if login_requerido():
        return redirect(url_for('login'))
    usuarios = facade.listar_usuarios()
    return render_template('listar_usuarios.html', usuarios=usuarios)


@app.route('/usuarios/crear', methods=['GET', 'POST'])
def crear_usuario():
    if login_requerido():
        return redirect(url_for('login'))
    if solo_admin():
        return render_template('sin_permiso.html')

    roles = facade.obtener_roles()

    if request.method == 'GET':
        return render_template('crear_usuario.html',
                               roles=roles, errores=[], exito=False,
                               datos=None, permisos=[])

    datos = {
        'nombre':   request.form.get('nombre', ''),
        'correo':   request.form.get('correo', ''),
        'password': request.form.get('password', ''),
        'rol_id':   request.form.get('rol_id', '')
    }

    resultado = facade.crear_usuario(datos)

    if resultado['exito']:
        return render_template('crear_usuario.html',
                               roles=roles, errores=[], exito=True,
                               datos=None, permisos=resultado['permisos'])

    return render_template('crear_usuario.html',
                           roles=roles, errores=resultado['errores'],
                           exito=False, datos=datos, permisos=[])


@app.route('/usuarios/<int:usuario_id>/editar-rol', methods=['GET', 'POST'])
def editar_rol(usuario_id):
    if login_requerido():
        return redirect(url_for('login'))
    if solo_admin():
        return render_template('sin_permiso.html')

    usuario = facade.obtener_usuario(usuario_id)
    if not usuario:
        return "Usuario no encontrado", 404

    roles = facade.obtener_roles()

    if request.method == 'GET':
        return render_template('editar_rol.html',
                               usuario=usuario, roles=roles,
                               errores=[], exito=False)

    rol_id    = request.form.get('rol_id', '')
    resultado = facade.editar_rol_usuario(usuario_id, rol_id)

    if resultado['exito']:
        usuario = facade.obtener_usuario(usuario_id)
        return render_template('editar_rol.html',
                               usuario=usuario, roles=roles,
                               errores=[], exito=True)

    return render_template('editar_rol.html',
                           usuario=usuario, roles=roles,
                           errores=resultado['errores'], exito=False)


# REPORTES
@app.route('/reportes')
def reportes():
    if login_requerido():
        return redirect(url_for('login'))
    if solo_admin():
        return render_template('sin_permiso.html')
    usuarios = facade.listar_usuarios()
    return render_template('reportes.html', usuarios=usuarios)


@app.route('/reportes/excel')
def reporte_excel():
    if login_requerido():
        return redirect(url_for('login'))
    if solo_admin():
        return render_template('sin_permiso.html')

    from flask import make_response
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io

    usuarios = facade.listar_usuarios()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    ws.merge_cells('A1:F1')
    ws['A1'] = 'REPORTE DE USUARIOS — GESTION DE VUELOS'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(fill_type='solid', fgColor='1a3a8f')
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['#', 'Nombre', 'Correo', 'Rol', 'Estado', 'Fecha de Registro']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid', fgColor='2d3748')
        cell.alignment = Alignment(horizontal='center')

    for row, u in enumerate(usuarios, 3):
        ws.cell(row=row, column=1, value=u['id'])
        ws.cell(row=row, column=2, value=u['nombre'])
        ws.cell(row=row, column=3, value=u['correo'])
        ws.cell(row=row, column=4, value=u['rol'].capitalize())
        ws.cell(row=row, column=5, value='Activo' if u['estado'] else 'Inactivo')
        fecha = u['fecha_creacion'].strftime('%d/%m/%Y') if u['fecha_creacion'] else '—'
        ws.cell(row=row, column=6, value=fecha)

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.read())
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_usuarios.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


@app.route('/reportes/pdf')
def reporte_pdf():
    if login_requerido():
        return redirect(url_for('login'))
    if solo_admin():
        return render_template('sin_permiso.html')

    from flask import make_response
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    import io

    usuarios = facade.listar_usuarios()
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    titulo = Paragraph('<b>REPORTE DE USUARIOS — GESTION DE VUELOS</b>', styles['Title'])
    elements.append(titulo)
    elements.append(Spacer(1, 20))

    data = [['#', 'Nombre', 'Correo', 'Rol', 'Estado', 'Fecha']]
    for u in usuarios:
        fecha = u['fecha_creacion'].strftime('%d/%m/%Y') if u['fecha_creacion'] else '—'
        data.append([str(u['id']), u['nombre'], u['correo'],
                     u['rol'].capitalize(), 'Activo' if u['estado'] else 'Inactivo', fecha])

    tabla = Table(data, colWidths=[30, 110, 150, 60, 60, 70])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a8f')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    elements.append(tabla)
    doc.build(elements)

    buffer.seek(0)
    response = make_response(buffer.read())
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_usuarios.pdf'
    response.headers['Content-Type'] = 'application/pdf'
    return response


if __name__ == '__main__':
    app.run(debug=True, port=5000)